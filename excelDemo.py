# -*-coding:utf-8-*-

import openpyxl, sys
reload(sys)


from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter

sys.setdefaultencoding('utf-8')

wb = Workbook()

dest_filename = 'test4.xlsx'

ws1 = wb.active
ws1.title = 'newDemo'

titleList = ['时间', '招聘企业', '学校', '地址']

for row in range(len(titleList)):
    c = row + 1
    ws1.cell(row=1, column=c, value=titleList[row])


tableValues = [['张学友', 15201062100, 18, '测试数据！'], ['李雷', 15201062598, 19, '测试数据！'],['Marry', 15201062191, 28, '测试数据！']]

for row in range(len(tableValues)):
    # ws1.append(tableValues[row])

    print tableValues[row]


ws2 = wb.create_sheet(title='Pi') #新建一个表
ws2['F5'] = 3.14 #给固定位置赋值

ws3 = wb.create_sheet(title='Data')#设置表名

# for row in range(1, 20):
#     for col in range(1, 20):
#         _ = ws3.cell(column=col, row = row, value = '{0}'.format(get_column_letter(col)))
#
#
# wb.save(filename=dest_filename)

